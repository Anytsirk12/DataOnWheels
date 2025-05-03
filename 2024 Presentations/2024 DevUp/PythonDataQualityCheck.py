

'''---------------README------------------
Full instructions available: https://dataonwheels.wordpress.com/?s=Power+BI%3A+Data+Quality+Checks+Using+Python+%26+SQL
The goal of this script is to compare DAX measures to corresponding SQL queries. 
To accomplish this, it uses an excel file containing authentication variables and query values.
This idea came from seeing the Execute Queries REST API in Power BI and wanting to use python to perform our data quality checks efficiently.
To connect to a Power BI Data Model, we need to pass an authentication token generated through an Azure AD App. 
To run a DAX query and SQL query, we need to loop through our excel file then put the queries into the API and SQL calls respectively.
Finally, this script takes those query results and puts it back into the excel file.
 
I have left some print() lines commented out, feel free to uncomment them to troubleshoot or walk through the code step-by-step.
 
You will need to swap out the excel_file variable with your excel file path. Other than that, the rest of the variables are managed inside your excel file. 
Once you have your variables and queries in excel, hit run and you're good to go. 
'''
 
#---------------------------------------#
#       Import libraries needed  
#---------------------------------------#
import requests
import adal
import json
import pymssql 
import pandas
import openpyxl
from win32com.client import Dispatch; import os
 
#---------------------------------------#
#    Build out authentication steps  
#---------------------------------------#
 
#----- Authentication Variables from Excel -----#
 
#Grab authentication data from our excel file
from openpyxl import load_workbook
#\\ escapes the \ so we can get the proper file path
excel_file = "C:\\Users\\krist\\Downloads\\PythonDataQualityChecker.xlsx"
wb = load_workbook(excel_file) 
sheets = wb.sheetnames
#print(sheets) #see the sheets in our workbook
cred_ws = wb['Credentials'] #access specific sheet called Credentials
cred_mapping = {} #this will help us build a dataframe from our table
for entry, data_boundary in cred_ws.tables.items():
    #parse the data within the ref boundary
    data = cred_ws[data_boundary]
    #extract the data 
    #the inner list comprehension gets the values for each cell in the table
    content = [[cell.value for cell in ent] #ent makes our code cleaner by handling nested data sources
               for ent in data
          ]
    header = content[0] #the contents excluding the header (aka column names)
    rest = content[1:] #create dataframe with the column names
    cred_df = pandas.DataFrame(rest, columns = header) #this is the dataframe we will use to get our creds
    cred_mapping[entry] = cred_df


#Use the dataframe to set up credential variables we can use later in the script
sqlserver = str(cred_df._get_value(0,"SQL_Server"))
sqldb = str(cred_df._get_value(0,"SQL_Database"))
sqluser = str(cred_df._get_value(0,"SQL_User"))
sqlpassword = str(cred_df._get_value(0,"SQL_Password"))
pbiclientid = str(cred_df._get_value(0,"PBI_ClientID"))
pbiclientsecret = str(cred_df._get_value(0,"PBI_ClientSecret"))
#pbiusername = str(cred_df._get_value(0,"PBI_Username"))
#pbipassword = str(cred_df._get_value(0,"PBI_Password"))
pbidatasetid = str(cred_df._get_value(0,"PBI_DatasetID"))
pbitenantid = str(cred_df._get_value(0,"PBI_TenantID"))
resource = 'https://analysis.windows.net/powerbi/api'
#check to make sure your variables are correct by uncommenting the next line
#print(sqlserver,sqldb,sqluser,sqlpassword,pbiclientid,pbiusername,pbipassword,pbidatasetid)
#print(sqlserver,sqldb)

#----- Power BI REST API Authentication -----#
 
authority_url = 'https://login.microsoftonline.com/{0}/oauth2/token'.format(pbitenantid)
data = {
    'grant_type': 'client_credentials',
    'client_id': pbiclientid,
    'client_secret': pbiclientsecret,
    'resource': resource
}
auth_response = requests.post(authority_url, data=data)
access_token = auth_response.json()['access_token'] 
pbiheader = {'Authorization': f'Bearer {access_token}'}
#try out the get workspaces REST API using our access token by uncommenting the next few lines
#get_workspaces_url = 'https://api.powerbi.com/v1.0/myorg/groups'
#r = requests.get(url=get_workspaces_url, headers=pbiheader)
#r.text will give us the response text for our get request, pretty neat!
#print(r.text)
#print(pbiheader)

#----- SQL Server Authentication -----#
 
try:
    sql_con = pymssql.connect(sqlserver,sqluser,sqlpassword,sqldb)
    sql_cursor = sql_con.cursor(as_dict=True)
except Exception as e:
    raise Exception(e)
 
 
#---------------------------------------#
#  Build out data quality check steps 
#---------------------------------------#
 
#----- Read excel to get quality check queries into a dataframe -----#
 
quality_ws = wb['Quality Check']
quality_mapping = {} #this will help us build a dataframe from our table
for entry, data_boundary in quality_ws.tables.items(): #grabs data dynamically from our table
    data = quality_ws[data_boundary] #parse the data within the ref boundary
    #the inner list comprehension gets the values for each cell in the table
    content = [[cell.value for cell in ent] #ent makes our code cleaner by handling nested data sources
               for ent in data
          ]
    header = content[0] #the contents excluding the header (aka column names)
    rest = content[1:] #create dataframe with the column names
    qualitydf = pandas.DataFrame(rest, columns = header) #this is the dataframe we will use to get our quality check queries
    quality_df = qualitydf.fillna(' ') #helps remove blank records from our excel file
    quality_mapping[entry] = quality_df
#print(quality_df)


 
#----- Open the specified excel file in edit mode -----#

xl = Dispatch("Excel.Application") #opens excel for us to edit
xl.Visible = True
edit_wb = xl.Workbooks.Open(excel_file)
edit_qc_sh = edit_wb.Worksheets("Quality Check")
 
#----- Set variables to use in our iterators -----#
 
qdf = quality_df.dropna() #removing any blank rows so we don't run blank sql queries lol
sqlrownumber = 1
pbirownumber = -1

#----- Run SQL queries and put results back into excel -----#
 
sql_queries = qdf.loc[:,"SQL_Query"] 
#print(sql_queries)


for query in sql_queries:
    sql_cursor.execute(str(query))
    #print (sql_cursor.fetchall()[0])
    for row in sql_cursor:
        #print(row) #this is our dictionary created from the sql query result
        key = list(row.items())[0] #gets the first item from the row dictionary
        sqlrownumber += 1 #puts our sql responses in the right rows
        #sql_result = key[1] #grabs just the result of our query
        sql_result = str(key[1]) #grabs just the result of our query #converting to string gets us all of the decimal places
        print(sql_result)
        quality_df.at[sqlrownumber,"SQL_Result"] = sql_result #this will put our results into the proper cell in our dataframe
        edit_qc_sh.Range(f"D{sqlrownumber}").Value = sql_result #this will put our results in the right excel cell
        #print(key[1]) #returns just our result values

print("SQL Succeeded")

#----- Run PBI DAX queries and put results back into excel -----#
 
pbi_queries = quality_df.loc[:,"PBI_DAX_Query"]
#print(pbi_queries)
execute_queries_url = f'https://api.powerbi.com/v1.0/myorg/datasets/{pbidatasetid}/executeQueries'


for items in pbi_queries:
    pbirownumber += 1
    pbiexcelrownumber = pbirownumber+2
    list_pbiquery = list(pbi_queries.items())[pbirownumber]
    #print(list_pbiquery)
    item_pbiquery = list_pbiquery[1]
    dax_query_json_body = {
                "queries":
                        [
                            {
                            "query": item_pbiquery
                            }
                        ]  
                }
    pbi_response = requests.post(url=execute_queries_url, headers=pbiheader,json=dax_query_json_body)
    query_json = pbi_response.text.encode().decode('utf-8-sig') #allows us to read the json response
    pbi_values = query_json.replace("}]}]}]}","") #drops the trailing encoding from json
    #print(query_json)
    pbi_result = pbi_values.split(":")[4] #grabs just the result of our query
    #print(json.dumps(pbi_result)) #grabs just the result of our query
    quality_df.at[pbiexcelrownumber,"PBI_Result"] = pbi_result #this will put our results into the proper cell in our dataframe
    edit_qc_sh.Range(f"C{pbiexcelrownumber}").Value = pbi_result #this will put our results in the right excel cell
    

print("PBI Succeeded")
 
#----- Save our changes back to the excel file -----#
 
edit_wb.Save() #saving our values back to excel

print("All done")
#'''